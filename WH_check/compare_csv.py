# -*- coding: utf-8 -*-
import csv
import pandas as pd
from datetime import datetime
import re
import os
import sys
from dateutil import parser as date_parser
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook


def clean_text(text):
    if not text or pd.isna(text):
        return ''
    text = str(text)
    text = re.sub(r'\s+', '', text)
    text = re.sub(r'[（(][^）)]*[）)]', '', text)
    text = text.replace('-', '').replace('\uff0c', ',').replace('\u3002', '.')
    text = text.replace('\u3001', ',').replace('\uff1b', ';').replace('\uff1a', ':')
    return text.strip()


def parse_date_input(date_str, default_year=2026):
    if not date_str or date_str.strip() == '':
        return None
    date_str = date_str.strip()
    if re.match(r'^\d+[./-]\d+$', date_str):
        parts = re.split(r'[./-]', date_str)
        if len(parts) == 2:
            month = int(parts[0])
            day = int(parts[1])
            try:
                dt = datetime(default_year, month, day)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                pass
    try:
        dt = date_parser.parse(date_str, fuzzy=True, default=datetime(default_year, 1, 1))
        if dt.year < 100:
            dt = dt.replace(year=default_year)
        return dt.strftime('%Y-%m-%d')
    except:
        pass
    return None


def get_date_range_from_user():
    print()
    print('请选择要对比的日期范围（默认年份为2026年）')
    print('='*60)
    print('支持格式示例：')
    print('  - 单日: 7.24 或 7/24')
    print('  - 范围: 7.22-7.26 或 07/22-07/26')
    print('  - 直接回车: 对比所有日期')
    print('  - 输入 q 退出')
    print('-'*60)
    while True:
        user_input = input('\n请输入日期: ').strip()
        if user_input.lower() in ['q', 'quit', 'exit']:
            print('已退出程序')
            sys.exit(0)
        if user_input == '':
            return None, None
        if '-' in user_input:
            parts = user_input.split('-')
            if len(parts) == 2:
                start_str = parts[0].strip()
                end_str = parts[1].strip()
                start_date = parse_date_input(start_str)
                end_date = parse_date_input(end_str)
                if start_date and end_date:
                    if start_date > end_date:
                        start_date, end_date = end_date, start_date
                    return start_date, end_date
                else:
                    print('日期格式无效，请重新输入')
                    continue
        single_date = parse_date_input(user_input)
        if single_date:
            return single_date, single_date
        else:
            print('日期格式无效，请重新输入')


def get_excel_sheets(excel_path):
    try:
        xl = pd.ExcelFile(excel_path)
        return xl.sheet_names
    except Exception as e:
        print('读取Excel文件失败:', e)
        return []


def select_sheet_interactive(sheets):
    for sheet in sheets:
        if 'PPG工业漆' in sheet:
            return sheet
    print()
    print('找到', len(sheets), '个子表，请选择：')
    for i, sheet in enumerate(sheets, 1):
        print(' ', i, '.', sheet)
    while True:
        try:
            choice = input('\n请输入子表编号 (1-{0}): '.format(len(sheets))).strip()
            if choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(sheets):
                    return sheets[idx]
            print('无效选择，请重新输入')
        except:
            print('无效输入，请重新输入')


def apply_excel_highlighting(wb):
    try:
        ws = wb['差异详情']
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        bold_font = Font(bold=True)
        diff_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and '差异_' in str(header):
                diff_cols.append(col)
        for col in diff_cols:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and '\u2717' in str(cell.value):
                    cell.fill = red_fill
                    cell.font = Font(color='FFFFFF', bold=True)
                    csv_col = col - 2
                    excel_col = col - 1
                    if csv_col >= 1:
                        ws.cell(row=row, column=csv_col).fill = yellow_fill
                        ws.cell(row=row, column=csv_col).font = bold_font
                    if excel_col >= 1:
                        ws.cell(row=row, column=excel_col).fill = yellow_fill
                        ws.cell(row=row, column=excel_col).font = bold_font
                elif cell.value and '\u2713' in str(cell.value):
                    cell.font = Font(color='00AA00', bold=True)
        for sheet_name in wb.sheetnames:
            ws_s = wb[sheet_name]
            for col in range(1, ws_s.max_column + 1):
                max_length = 0
                for row in range(1, min(ws_s.max_row + 1, 100)):
                    cell_value = ws_s.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max_length + 2, 50)
                ws_s.column_dimensions[ws_s.cell(row=1, column=col).column_letter].width = adjusted_width
            ws_s.freeze_panes = 'A2'
    except Exception as e:
        print('高亮失败:', e)


def read_csv_file(csv_file_path):
    """读取CSV文件，返回DataFrame"""
    if not os.path.exists(csv_file_path):
        print('CSV文件不存在:', csv_file_path)
        return None

    try:
        # 尝试用 utf-8-sig 读取（带BOM的UTF-8）
        try:
            df = pd.read_csv(csv_file_path, encoding='utf-8-sig', dtype=str)
        except:
            df = pd.read_csv(csv_file_path, encoding='gbk', dtype=str)

        print('CSV:', len(df), '条')
        return df
    except Exception as e:
        print('读取CSV失败:', e)
        return None


def compare_csv_with_excel(csv_file_path, excel_file_path,
                           sheet_name=None, start_date=None, end_date=None):
    if not os.path.exists(csv_file_path):
        print('CSV文件不存在:', csv_file_path)
        return
    if not os.path.exists(excel_file_path):
        print('Excel文件不存在:', excel_file_path)
        return

    if start_date is None and end_date is None:
        start_date, end_date = get_date_range_from_user()
    if sheet_name is None:
        sheets = get_excel_sheets(excel_file_path)
        if not sheets:
            return
        sheet_name = select_sheet_interactive(sheets)

    print()
    print('子表:', sheet_name)
    if start_date:
        print('日期:', start_date, '至', end_date)
    else:
        print('日期: 全部')

    # === 读取CSV ===
    df_csv = read_csv_file(csv_file_path)
    if df_csv is None:
        return

    # === 读取Excel台账 ===
    try:
        df_excel = pd.read_excel(excel_file_path, sheet_name=sheet_name, dtype=str)
        print('Excel台账:', len(df_excel), '条')
    except Exception as e:
        print('读取Excel失败:', e)
        return

    # 检查必需列
    required_columns = ['单号', '下单日期', '始发城市']
    missing_cols = [col for col in required_columns if col not in df_excel.columns]
    if missing_cols:
        print('Excel缺少必要列:', missing_cols)
        return

    # CSV也必须有单号和下单日期
    csv_required = ['单号', '下单日期']
    csv_missing = [col for col in csv_required if col not in df_csv.columns]
    if csv_missing:
        print('CSV缺少必要列:', csv_missing)
        print('CSV列名:', list(df_csv.columns))
        return

    # === 筛选日期范围（Excel台账） ===
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
            df_excel['下单日期_dt'] = pd.to_datetime(df_excel['下单日期'], errors='coerce')
            df_excel = df_excel[(df_excel['下单日期_dt'] >= start) & (df_excel['下单日期_dt'] <= end)]
        except Exception as e:
            print('日期筛选失败:', e)

    # === 只保留马鞍山库的订单（Excel台账） ===
    df_excel_ma = df_excel[df_excel['始发城市'] == '马鞍山库'].copy()
    print('Excel台账（仅马鞍山库）:', len(df_excel_ma), '条')



    # === CSV同样只保留马鞍山库的订单（与台账口径一致） ===
    # 注意：这里不再按“单号是否存在于台账”来预过滤CSV，
    # 否则CSV中台账没有的订单（即CSV多出的订单）会被静默丢弃，
    # 导致“CSV多出订单号”永远为空。
    if '始发城市' in df_csv.columns:
        df_csv['始发城市_str'] = df_csv['始发城市'].astype(str).str.strip()
        csv_filtered = df_csv[df_csv['始发城市_str'] == '马鞍山库'].copy()
    else:
        csv_filtered = df_csv.copy()

    # 进一步按日期筛选CSV
    if start_date and end_date:
        try:
            csv_filtered['下单日期_dt'] = pd.to_datetime(csv_filtered['下单日期'], errors='coerce')
            csv_filtered = csv_filtered[(csv_filtered['下单日期_dt'] >= start) & (csv_filtered['下单日期_dt'] <= end)]
        except Exception as e:
            print('CSV日期筛选失败:', e)

    print('CSV（马鞍山库）:', len(csv_filtered), '条')

    # === 构建字典 ===
    excel_dict = {}
    for _, row in df_excel_ma.iterrows():
        order_no = str(row['单号']) if pd.notna(row['单号']) else ''
        if not order_no or order_no == 'nan':
            continue
        excel_dict[order_no] = {
            'order_date': str(row['下单日期']) if '下单日期' in df_excel_ma.columns and pd.notna(row['下单日期']) else '',
            'receiver': str(row['收货单位']) if '收货单位' in df_excel_ma.columns and pd.notna(row['收货单位']) else '',
            'address': str(row['收货地址']) if '收货地址' in df_excel_ma.columns and pd.notna(row['收货地址']) else '',
            'contact': str(row['收货人']) if '收货人' in df_excel_ma.columns and pd.notna(row['收货人']) else '',
            'requirement': str(row['客户要求']) if '客户要求' in df_excel_ma.columns and pd.notna(row['客户要求']) else '',
            'quantity': str(row['数量']) if '数量' in df_excel_ma.columns and pd.notna(row['数量']) else '',
            'weight': str(row['重量']) if '重量' in df_excel_ma.columns and pd.notna(row['重量']) else '',
            'shipping': str(row['发运方式']) if '发运方式' in df_excel_ma.columns and pd.notna(row['发运方式']) else '',
            'city': str(row['到货城市']) if '到货城市' in df_excel_ma.columns and pd.notna(row['到货城市']) else '',
            'province': str(row['到货省份']) if '到货省份' in df_excel_ma.columns and pd.notna(row['到货省份']) else '',
            'product': str(row['产品特性']) if '产品特性' in df_excel_ma.columns and pd.notna(row['产品特性']) else '',
        }

    csv_dict = {}
    for _, row in csv_filtered.iterrows():
        order_no = str(row['单号']) if pd.notna(row['单号']) else ''
        if not order_no or order_no == 'nan':
            continue
        csv_dict[order_no] = {
            'order_date': str(row['下单日期']) if pd.notna(row['下单日期']) else '',
            'receiver': str(row['收货单位']) if '收货单位' in row and pd.notna(row['收货单位']) else '',
            'address': str(row['收货地址']) if '收货地址' in row and pd.notna(row['收货地址']) else '',
            'contact': str(row['收货人']) if '收货人' in row and pd.notna(row['收货人']) else '',
            'requirement': str(row['客户要求']) if '客户要求' in row and pd.notna(row['客户要求']) else '',
            'quantity': str(row['数量']) if '数量' in row and pd.notna(row['数量']) else '',
            'weight': str(row['重量']) if '重量' in row and pd.notna(row['重量']) else '',
            'shipping': str(row['发运方式']) if '发运方式' in row and pd.notna(row['发运方式']) else '',
            'city': str(row['到货城市']) if '到货城市' in row and pd.notna(row['到货城市']) else '',
            'province': str(row['到货省份']) if '到货省份' in row and pd.notna(row['到货省份']) else '',
            'product': str(row['产品特性']) if '产品特性' in row and pd.notna(row['产品特性']) else '',
        }

    # === 按天分组统计（每日统计） ===
    df_excel_ma['下单日期_dt'] = pd.to_datetime(df_excel_ma['下单日期'], errors='coerce')
    all_dates = sorted(df_excel_ma['下单日期_dt'].dropna().dt.date.unique())

    daily_stats = []

    for date_obj in all_dates:
        date_str = date_obj.strftime('%Y-%m-%d')

        excel_day_orders = set(
            str(o) for o in df_excel_ma[df_excel_ma['下单日期_dt'].dt.date == date_obj]['单号']
            if pd.notna(o)
        )

        csv_day_orders = set()
        for oid, oitem in csv_dict.items():
            od = oitem.get('order_date', '')
            if od:
                try:
                    if '/' in od:
                        parts = od.split('/')
                        if len(parts) == 3:
                            dt = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                            if dt.date() == date_obj:
                                csv_day_orders.add(oid)
                    elif '-' in od:
                        dt = datetime.strptime(od, '%Y-%m-%d')
                        if dt.date() == date_obj:
                            csv_day_orders.add(oid)
                except:
                    pass

        common_day = excel_day_orders & csv_day_orders
        excel_extra = excel_day_orders - csv_day_orders
        csv_extra = csv_day_orders - excel_day_orders

        diff_orders = set()
        for oid in common_day:
            c = csv_dict[oid]
            e = excel_dict[oid]
            for field in ['receiver', 'address', 'contact', 'requirement', 'quantity', 'weight', 'shipping', 'city', 'province', 'product']:
                c_val = c.get(field, '').strip()
                e_val = e.get(field, '').strip()
                if field in ['quantity', 'weight']:
                    c_num = re.sub(r'[^\d.]', '', c_val)
                    e_num = re.sub(r'[^\d.]', '', e_val)
                    if c_num != e_num:
                        diff_orders.add(oid)
                        break
                else:
                    c_clean = clean_text(c_val)
                    e_clean = clean_text(e_val)
                    if c_clean != e_clean:
                        diff_orders.add(oid)
                        break

        row = {
            '日期': date_str,
            '台账数量': len(excel_day_orders),
            'CSV数量': len(csv_day_orders),
            '共同数量': len(common_day),
            '差异数量': len(diff_orders),
            '台账多出订单号': '\u3001'.join(sorted(excel_extra)) if excel_extra else '',
            'CSV多出订单号': '\u3001'.join(sorted(csv_extra)) if csv_extra else '',
            '差异订单号': '\u3001'.join(sorted(diff_orders)) if diff_orders else '',
        }
        daily_stats.append(row)

    # === 订单级差异详情 ===
    csv_orders = set(csv_dict.keys())
    excel_orders = set(excel_dict.keys())
    common_orders = csv_orders & excel_orders

    field_names = {
        'receiver': '收货单位', 'address': '收货地址', 'contact': '收货人',
        'requirement': '客户要求', 'quantity': '数量', 'weight': '重量',
        'shipping': '发运方式', 'city': '到货城市', 'province': '到货省份', 'product': '产品特性'
    }

    diff_data = []
    field_diff_count = {f: 0 for f in field_names.values()}

    # 按日期排序
    def get_order_date(no):
        item = csv_dict.get(no, excel_dict.get(no, {}))
        return item.get('order_date', '')
    sorted_orders = sorted(common_orders, key=get_order_date)

    for order_no in sorted_orders:
        csv_item = csv_dict[order_no]
        excel_item = excel_dict[order_no]
        order_date = csv_item.get('order_date', '') or excel_item.get('order_date', '')
        row = {'日期': order_date, '订单号': order_no}
        has_diff = False
        for field, name in field_names.items():
            c_val = csv_item.get(field, '').strip()
            e_val = excel_item.get(field, '').strip()
            row['CSV_' + name] = c_val
            row['台账_' + name] = e_val
            if field in ['quantity', 'weight']:
                c_num = re.sub(r'[^\d.]', '', c_val)
                e_num = re.sub(r'[^\d.]', '', e_val)
                if c_num != e_num:
                    row['差异_' + name] = '\u2717'
                    field_diff_count[name] += 1
                    has_diff = True
                else:
                    row['差异_' + name] = '\u2713'
            else:
                c_clean = clean_text(c_val)
                e_clean = clean_text(e_val)
                if c_clean != e_clean:
                    row['差异_' + name] = '\u2717'
                    field_diff_count[name] += 1
                    has_diff = True
                else:
                    row['差异_' + name] = '\u2713'
        if has_diff:
            diff_data.append(row)

    # === 写入Excel（两个子表） ===
    date_suffix = ''
    if start_date and end_date:
        if start_date == end_date:
            date_suffix = '_' + start_date
        else:
            date_suffix = '_' + start_date + '_至_' + end_date

    output_file = 'CSV差异报告' + date_suffix + '.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 子表1: 每日统计
        df_daily = pd.DataFrame(daily_stats)
        if not df_daily.empty:
            total_row = {
                '日期': '合计',
                '台账数量': int(df_daily['台账数量'].sum()),
                'CSV数量': int(df_daily['CSV数量'].sum()),
                '共同数量': int(df_daily['共同数量'].sum()),
                '差异数量': int(df_daily['差异数量'].sum()),
                '台账多出订单号': '',
                'CSV多出订单号': '',
                '差异订单号': '',
            }
            df_daily = pd.concat([df_daily, pd.DataFrame([total_row])], ignore_index=True)
        df_daily.to_excel(writer, sheet_name='每日统计', index=False)

        # 子表2: 差异详情
        if diff_data:
            df_out = pd.DataFrame(diff_data)
            stats_row = {'日期': '', '订单号': '=== 差异统计 ==='}
            for name in field_names.values():
                stats_row['CSV_' + name] = ''
                stats_row['台账_' + name] = ''
                stats_row['差异_' + name] = int(field_diff_count[name]) if field_diff_count[name] > 0 else ''
            df_out = pd.concat([df_out, pd.DataFrame([stats_row])], ignore_index=True)
        else:
            df_out = pd.DataFrame([{'日期': '', '订单号': '所有订单完全一致，无差异'}])
        df_out.to_excel(writer, sheet_name='差异详情', index=False)

    try:
        wb = load_workbook(output_file)
        apply_excel_highlighting(wb)
        wb.save(output_file)
    except Exception as e:
        print('高亮失败:', e)

    # === 精简输出 ===
    print()
    print('已导出:', output_file)
    total_diff = sum(field_diff_count.values())
    print('差异订单:', len(diff_data), '个 | 差异字段:', total_diff, '处')
    diff_fields = [(n, c) for n, c in field_diff_count.items() if c > 0]
    if diff_fields:
        for n, c in diff_fields:
            print(' .', n, ':', c, '处')

    print()
    print('每日统计:')
    for ds in daily_stats:
        extras = []
        if ds['台账多出订单号']:
            extras.append('台账多' + str(ds['台账数量'] - ds['共同数量']))
        if ds['CSV多出订单号']:
            extras.append('CSV多' + str(ds['CSV数量'] - ds['共同数量']))
        extra_str = ''
        if extras:
            extra_str = ' (' + ', '.join(extras) + ')'
        print('  ', ds['日期'], ': 台账', ds['台账数量'], ' CSV', ds['CSV数量'], ' 共同', ds['共同数量'], ' 差异', ds['差异数量'], extra_str)


if __name__ == '__main__':
    csv_path = r'D:\project\WH_check\file\自动接单表_PPG 芜湖(工业漆)_总接单表.csv'
    excel_path = r'D:\project\WH_check\file\调度台账 2023年6月.xlsx'

    print()
    print('='*60)
    print('PPG芜湖工业漆 CSV vs 台账 对比工具（马鞍山库）')
    print('='*60)
    print('说明：CSV为自动接单表导出数据，与调度台账进行比对')
    print('比对字段：单号、下单日期、收货单位、收货地址、收货人、')
    print('         客户要求、数量、重量、发运方式、到货城市、')
    print('         到货省份、产品特性')
    print('='*60)

    if not os.path.exists(csv_path):
        print('找不到CSV文件:', csv_path)
        csv_path = input('请输入CSV文件路径: ').strip()
    if not os.path.exists(excel_path):
        print('找不到Excel文件:', excel_path)
        excel_path = input('请输入Excel文件路径: ').strip()

    compare_csv_with_excel(
        csv_file_path=csv_path,
        excel_file_path=excel_path,
        sheet_name=None,
        start_date=None,
        end_date=None
    )
