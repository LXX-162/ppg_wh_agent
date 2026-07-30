# -*- coding: utf-8 -*-
import json
import pandas as pd
from datetime import datetime
import re
import os
import sys
from dateutil import parser as date_parser
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook


def clean_text(text):
    if not text or pd.isna(text):
        return ''
    text = str(text)
    text = re.sub(r'\s+', '', text)
    text = re.sub(r'[（(][^）)]*[）)]', '', text)
    text = text.replace('-', '').replace('\uff0c', ',').replace('\u3002', '.')
    text = text.replace('\u3001', ',').replace('\uff1b', ';').replace('\uff1a', ':')
    return text.strip()


def parse_date_input(date_str, default_year=2026):
    if not date_str or date_str.strip() == '':
        return None
    date_str = date_str.strip()
    if re.match(r'^\d+[./-]\d+$', date_str):
        parts = re.split(r'[./-]', date_str)
        if len(parts) == 2:
            month = int(parts[0])
            day = int(parts[1])
            try:
                dt = datetime(default_year, month, day)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                pass
    try:
        dt = date_parser.parse(date_str, fuzzy=True, default=datetime(default_year, 1, 1))
        if dt.year < 100:
            dt = dt.replace(year=default_year)
        return dt.strftime('%Y-%m-%d')
    except:
        pass
    return None


def get_date_range_from_user():
    print()
    print('='*60)
    print('请选择要对比的日期范围（默认年份为2026年）')
    print('='*60)
    print('支持格式示例：')
    print('  - 单日: 7.24 或 7/24')
    print('  - 范围: 7.22-7.26 或 07/22-07/26')
    print('  - 直接回车: 对比所有日期')
    print('  - 输入 q 退出')
    print('-'*60)
    while True:
        user_input = input('\n请输入日期: ').strip()
        if user_input.lower() in ['q', 'quit', 'exit']:
            print('已退出程序')
            sys.exit(0)
        if user_input == '':
            return None, None
        if '-' in user_input:
            parts = user_input.split('-')
            if len(parts) == 2:
                start_str = parts[0].strip()
                end_str = parts[1].strip()
                start_date = parse_date_input(start_str)
                end_date = parse_date_input(end_str)
                if start_date and end_date:
                    if start_date > end_date:
                        start_date, end_date = end_date, start_date
                    return start_date, end_date
                else:
                    print('日期格式无效，请重新输入')
                    continue
        single_date = parse_date_input(user_input)
        if single_date:
            return single_date, single_date
        else:
            print('日期格式无效，请重新输入')


def get_excel_sheets(excel_path):
    try:
        xl = pd.ExcelFile(excel_path)
        return xl.sheet_names
    except Exception as e:
        print('读取Excel文件失败:', e)
        return []


def select_sheet_interactive(sheets):
    for sheet in sheets:
        if 'PPG工业漆' in sheet:
            return sheet
    print()
    print('找到', len(sheets), '个子表，请选择：')
    for i, sheet in enumerate(sheets, 1):
        print(' ', i, '.', sheet)
    while True:
        try:
            choice = input('\n请输入子表编号 (1-{0}): '.format(len(sheets))).strip()
            if choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(sheets):
                    return sheets[idx]
            print('无效选择，请重新输入')
        except:
            print('无效输入，请重新输入')


def apply_excel_highlighting(wb):
    try:
        ws = wb['差异详情']
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        bold_font = Font(bold=True)
        diff_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and '差异_' in str(header):
                diff_cols.append(col)
        for col in diff_cols:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and '\u2717' in str(cell.value):
                    cell.fill = red_fill
                    cell.font = Font(color='FFFFFF', bold=True)
                    json_col = col - 2
                    excel_col = col - 1
                    if json_col >= 1:
                        ws.cell(row=row, column=json_col).fill = yellow_fill
                        ws.cell(row=row, column=json_col).font = bold_font
                    if excel_col >= 1:
                        ws.cell(row=row, column=excel_col).fill = yellow_fill
                        ws.cell(row=row, column=excel_col).font = bold_font
                elif cell.value and '\u2713' in str(cell.value):
                    cell.font = Font(color='00AA00', bold=True)
        for sheet_name in wb.sheetnames:
            ws_s = wb[sheet_name]
            for col in range(1, ws_s.max_column + 1):
                max_length = 0
                for row in range(1, min(ws_s.max_row + 1, 100)):
                    cell_value = ws_s.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max_length + 2, 50)
                ws_s.column_dimensions[ws_s.cell(row=1, column=col).column_letter].width = adjusted_width
            ws_s.freeze_panes = 'A2'
    except Exception as e:
        print('高亮失败:', e)


def compare_orders_by_date_range(json_file_path, excel_file_path,
                                 sheet_name=None, start_date=None, end_date=None):
    if not os.path.exists(json_file_path):
        print('JSON文件不存在:', json_file_path)
        return
    if not os.path.exists(excel_file_path):
        print('Excel文件不存在:', excel_file_path)
        return

    if start_date is None and end_date is None:
        start_date, end_date = get_date_range_from_user()
    if sheet_name is None:
        sheets = get_excel_sheets(excel_file_path)
        if not sheets:
            return
        sheet_name = select_sheet_interactive(sheets)

    print()
    print('子表:', sheet_name)
    if start_date:
        print('日期:', start_date, '至', end_date)
    else:
        print('日期: 全部')

    # === 读取JSON ===
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        print('JSON:', len(json_data), '条')
    except Exception as e:
        print('读取JSON失败:', e)
        return

    if isinstance(json_data, dict):
        json_list = []
        for key, value in json_data.items():
            if isinstance(value, dict):
                if 'order_no' not in value:
                    value['order_no'] = key
                json_list.append(value)
        json_data = json_list

    # === 读取Excel ===
    try:
        df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        print('Excel:', len(df), '条')
    except Exception as e:
        print('读取Excel失败:', e)
        return

    required_columns = ['单号', '下单日期', '始发城市']
    missing_cols = [col for col in required_columns if col not in df.columns]
    if missing_cols:
        print('Excel缺少必要列:', missing_cols)
        return

    # === 筛选日期 ===
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
            df['下单日期_dt'] = pd.to_datetime(df['下单日期'], errors='coerce')
            df = df[(df['下单日期_dt'] >= start) & (df['下单日期_dt'] <= end)]
        except Exception as e:
            print('日期筛选失败:', e)

    # === 只保留马鞍山库的订单 ===
    df_ma = df[df['始发城市'] == '马鞍山库'].copy()
    print('仅马鞍山库:', len(df_ma), '条')

    ma_order_nos = set(str(o) for o in df_ma['单号'] if pd.notna(o))

    json_filtered = []
    for item in json_data:
        if not isinstance(item, dict):
            continue
        order_no = item.get('order_no', '')
        if order_no and order_no in ma_order_nos:
            if start_date and end_date:
                order_date = item.get('order_date', '')
                if order_date:
                    try:
                        if '/' in order_date:
                            parts = order_date.split('/')
                            if len(parts) == 3:
                                dt = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                                if start <= dt <= end:
                                    json_filtered.append(item)
                        elif '-' in order_date:
                            dt = datetime.strptime(order_date, '%Y-%m-%d')
                            if start <= dt <= end:
                                json_filtered.append(item)
                        else:
                            json_filtered.append(item)
                    except:
                        json_filtered.append(item)
                else:
                    json_filtered.append(item)
            else:
                json_filtered.append(item)

    # === 构建字典 ===
    excel_dict = {}
    for _, row in df_ma.iterrows():
        order_no = str(row['单号']) if pd.notna(row['单号']) else ''
        if not order_no or order_no == 'nan':
            continue
        excel_dict[order_no] = {
            'receiver': str(row['收货单位']) if '收货单位' in df.columns and pd.notna(row['收货单位']) else '',
            'address': str(row['收货地址']) if '收货地址' in df.columns and pd.notna(row['收货地址']) else '',
            'contact': str(row['收货人']) if '收货人' in df.columns and pd.notna(row['收货人']) else '',
            'requirement': str(row['客户要求']) if '客户要求' in df.columns and pd.notna(row['客户要求']) else '',
            'quantity': str(row['数量']) if '数量' in df.columns and pd.notna(row['数量']) else '',
            'weight': str(row['重量']) if '重量' in df.columns and pd.notna(row['重量']) else '',
            'shipping': str(row['发运方式']) if '发运方式' in df.columns and pd.notna(row['发运方式']) else '',
            'city': str(row['到货城市']) if '到货城市' in df.columns and pd.notna(row['到货城市']) else '',
            'province': str(row['到货省份']) if '到货省份' in df.columns and pd.notna(row['到货省份']) else '',
            'product': str(row['产品特性']) if '产品特性' in df.columns and pd.notna(row['产品特性']) else '',
            'order_date': str(row['下单日期']) if '下单日期' in df.columns and pd.notna(row['下单日期']) else '',
        }

    json_dict = {}
    for item in json_filtered:
        if not isinstance(item, dict):
            continue
        order_no = item.get('order_no', '')
        if not order_no:
            continue
        json_dict[order_no] = {
            'receiver': item.get('receiver', ''),
            'address': item.get('address', ''),
            'contact': item.get('contact', ''),
            'requirement': item.get('requirement', ''),
            'quantity': item.get('quantity', ''),
            'weight': item.get('weight', ''),
            'shipping': item.get('发运方式', ''),
            'city': item.get('到货城市', ''),
            'province': item.get('到货省份', ''),
            'product': item.get('危险品类别', ''),
            'order_date': item.get('order_date', ''),
        }

    # === 按天分组统计（每日统计） ===
    df_ma['下单日期_dt'] = pd.to_datetime(df_ma['下单日期'], errors='coerce')
    all_dates = sorted(df_ma['下单日期_dt'].dropna().dt.date.unique())

    daily_stats = []

    for date_obj in all_dates:
        date_str = date_obj.strftime('%Y-%m-%d')

        excel_day_orders = set(
            str(o) for o in df_ma[df_ma['下单日期_dt'].dt.date == date_obj]['单号']
            if pd.notna(o)
        )

        json_day_orders = set()
        for oid, oitem in json_dict.items():
            od = oitem.get('order_date', '')
            if od:
                try:
                    if '/' in od:
                        parts = od.split('/')
                        if len(parts) == 3:
                            dt = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                            if dt.date() == date_obj:
                                json_day_orders.add(oid)
                    elif '-' in od:
                        dt = datetime.strptime(od, '%Y-%m-%d')
                        if dt.date() == date_obj:
                            json_day_orders.add(oid)
                except:
                    pass

        common_day = excel_day_orders & json_day_orders
        excel_extra = excel_day_orders - json_day_orders
        json_extra = json_day_orders - excel_day_orders

        diff_orders = set()
        for oid in common_day:
            j = json_dict[oid]
            e = excel_dict[oid]
            for field in ['receiver', 'address', 'contact', 'requirement', 'quantity', 'weight', 'shipping', 'city', 'province', 'product']:
                j_val = j.get(field, '').strip()
                e_val = e.get(field, '').strip()
                if field in ['quantity', 'weight']:
                    j_num = re.sub(r'[^\d.]', '', j_val)
                    e_num = re.sub(r'[^\d.]', '', e_val)
                    if j_num != e_num:
                        diff_orders.add(oid)
                        break
                else:
                    j_clean = clean_text(j_val)
                    e_clean = clean_text(e_val)
                    if j_clean != e_clean:
                        diff_orders.add(oid)
                        break

        row = {
            '日期': date_str,
            '台账数量': len(excel_day_orders),
            'JSON数量': len(json_day_orders),
            '共同数量': len(common_day),
            '差异数量': len(diff_orders),
            '台账多出订单号': '\u3001'.join(sorted(excel_extra)) if excel_extra else '',
            'JSON多出订单号': '\u3001'.join(sorted(json_extra)) if json_extra else '',
            '差异订单号': '\u3001'.join(sorted(diff_orders)) if diff_orders else '',
        }
        daily_stats.append(row)

    # === 订单级差异详情 ===
    json_orders = set(json_dict.keys())
    excel_orders = set(excel_dict.keys())
    common_orders = json_orders & excel_orders

    field_names = {
        'receiver': '收货单位', 'address': '收货地址', 'contact': '收货人',
        'requirement': '客户要求', 'quantity': '数量', 'weight': '重量',
        'shipping': '发运方式', 'city': '到货城市', 'province': '到货省份', 'product': '产品特性'
    }

    diff_data = []
    field_diff_count = {f: 0 for f in field_names.values()}

    # 按日期排序
    def get_order_date(no):
        item = json_dict.get(no, excel_dict.get(no, {}))
        return item.get('order_date', '')
    sorted_orders = sorted(common_orders, key=get_order_date)

    for order_no in sorted_orders:
        json_item = json_dict[order_no]
        excel_item = excel_dict[order_no]
        order_date = json_item.get('order_date', '') or excel_item.get('order_date', '')
        row = {'日期': order_date, '订单号': order_no}
        has_diff = False
        for field, name in field_names.items():
            j_val = json_item.get(field, '').strip()
            e_val = excel_item.get(field, '').strip()
            row['JSON_' + name] = j_val
            row['台账_' + name] = e_val
            if field in ['quantity', 'weight']:
                j_num = re.sub(r'[^\d.]', '', j_val)
                e_num = re.sub(r'[^\d.]', '', e_val)
                if j_num != e_num:
                    row['差异_' + name] = '\u2717'
                    field_diff_count[name] += 1
                    has_diff = True
                else:
                    row['差异_' + name] = '\u2713'
            else:
                j_clean = clean_text(j_val)
                e_clean = clean_text(e_val)
                if j_clean != e_clean:
                    row['差异_' + name] = '\u2717'
                    field_diff_count[name] += 1
                    has_diff = True
                else:
                    row['差异_' + name] = '\u2713'
        if has_diff:
            diff_data.append(row)

    # === 写入Excel（两个子表） ===
    date_suffix = ''
    if start_date and end_date:
        if start_date == end_date:
            date_suffix = '_' + start_date
        else:
            date_suffix = '_' + start_date + '_至_' + end_date

    output_file = '差异报告' + date_suffix + '.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 子表1: 每日统计
        df_daily = pd.DataFrame(daily_stats)
        if not df_daily.empty:
            total_row = {
                '日期': '合计',
                '台账数量': int(df_daily['台账数量'].sum()),
                'JSON数量': int(df_daily['JSON数量'].sum()),
                '共同数量': int(df_daily['共同数量'].sum()),
                '差异数量': int(df_daily['差异数量'].sum()),
                '台账多出订单号': '',
                'JSON多出订单号': '',
                '差异订单号': '',
            }
            df_daily = pd.concat([df_daily, pd.DataFrame([total_row])], ignore_index=True)
        df_daily.to_excel(writer, sheet_name='每日统计', index=False)

        # 子表2: 差异详情
        if diff_data:
            df_out = pd.DataFrame(diff_data)
            stats_row = {'日期': '', '订单号': '=== 差异统计 ==='}
            for name in field_names.values():
                stats_row['JSON_' + name] = ''
                stats_row['台账_' + name] = ''
                stats_row['差异_' + name] = int(field_diff_count[name]) if field_diff_count[name] > 0 else ''
            df_out = pd.concat([df_out, pd.DataFrame([stats_row])], ignore_index=True)
        else:
            df_out = pd.DataFrame([{'日期': '', '订单号': '所有订单完全一致，无差异'}])
        df_out.to_excel(writer, sheet_name='差异详情', index=False)

    try:
        wb = load_workbook(output_file)
        apply_excel_highlighting(wb)
        wb.save(output_file)
    except Exception as e:
        print('高亮失败:', e)

    # === 精简输出 ===
    print()
    print('已导出:', output_file)
    total_diff = sum(field_diff_count.values())
    print('差异订单:', len(diff_data), '个 | 差异字段:', total_diff, '处')
    diff_fields = [(n, c) for n, c in field_diff_count.items() if c > 0]
    if diff_fields:
        for n, c in diff_fields:
            print(' .', n, ':', c, '处')

    print()
    print('每日统计:')
    for ds in daily_stats:
        extras = []
        if ds['台账多出订单号']:
            extras.append('台账多' + str(ds['台账数量'] - ds['共同数量']))
        if ds['JSON多出订单号']:
            extras.append('JSON多' + str(ds['JSON数量'] - ds['共同数量']))
        extra_str = ''
        if extras:
            extra_str = ' (' + ', '.join(extras) + ')'
        print('  ', ds['日期'], ': 台账', ds['台账数量'], ' JSON', ds['JSON数量'], ' 共同', ds['共同数量'], ' 差异', ds['差异数量'], extra_str)


if __name__ == '__main__':
    json_path = r'D:\project\ppg_wh_agent\output\cache\pending_orders.json'
    excel_path = r'D:\project\WH_check\file\调度台账 2023年6月.xlsx'

    print()
    print('='*60)
    print('PPG芜湖工业漆 订单对比工具（马鞍山库）')
    print('='*60)

    if not os.path.exists(json_path):
        print('找不到JSON文件:', json_path)
        json_path = input('请输入JSON文件路径: ').strip()
    if not os.path.exists(excel_path):
        print('找不到Excel文件:', excel_path)
        excel_path = input('请输入Excel文件路径: ').strip()

    compare_orders_by_date_range(
        json_file_path=json_path,
        excel_file_path=excel_path,
        sheet_name=None,
        start_date=None,
        end_date=None
    )
