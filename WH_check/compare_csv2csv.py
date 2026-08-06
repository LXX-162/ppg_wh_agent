# -*- coding: utf-8 -*-
"""
compare_csv2csv.py  —— PPG芜湖工业漆 两个CSV互相比对

对比两个格式完全相同的CSV文件（如 总接单表 与 副本_总接单表），
逐订单逐字段比对差异，并按天分组统计每天差异。

用法：
    python compare_csv2csv.py
    （脚本会提示选择日期范围，直接回车=全部日期，输入 q 退出）

输出：
    CSV2CSV差异报告_YYYY-MM-DD_至_YYYY-MM-DD.xlsx
"""
import csv
import re
import os
import sys
import pandas as pd
from datetime import datetime
from dateutil import parser as date_parser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


# ============ 通用工具（沿用 compare_csv.py 逻辑） ============

def clean_text(text):
    """清洗文本：去空白、括号内容、连字符，统一全角标点为半角"""
    if not text or pd.isna(text):
        return ''
    text = str(text)
    text = re.sub(r'\s+', '', text)
    text = re.sub(r'[（(][^）)]*[）)]', '', text)
    text = text.replace('-', '').replace('\uff0c', ',').replace('\u3002', '.')
    text = text.replace('\u3001', ',').replace('\uff1b', ';').replace('\uff1a', ':')
    return text.strip()


def parse_date_input(date_str, default_year=2026):
    if not date_str or date_str.strip() == '':
        return None
    date_str = date_str.strip()
    if re.match(r'^\d+[./-]\d+$', date_str):
        parts = re.split(r'[./-]', date_str)
        if len(parts) == 2:
            month = int(parts[0])
            day = int(parts[1])
            try:
                dt = datetime(default_year, month, day)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                pass
    try:
        dt = date_parser.parse(date_str, fuzzy=True, default=datetime(default_year, 1, 1))
        if dt.year < 100:
            dt = dt.replace(year=default_year)
        return dt.strftime('%Y-%m-%d')
    except:
        return None


def get_date_range_from_user():
    print()
    print('请选择要对比的日期范围（默认年份为2026年）')
    print('=' * 60)
    print('支持格式示例：')
    print('  - 单日: 7.24 或 7/24')
    print('  - 范围: 7.22-7.26 或 07/22-07/26')
    print('  - 直接回车: 对比所有日期')
    print('  - 输入 q 退出')
    print('-' * 60)
    while True:
        user_input = input('\n请输入日期: ').strip()
        if user_input.lower() in ['q', 'quit', 'exit']:
            print('已退出程序')
            sys.exit(0)
        if user_input == '':
            return None, None
        if '-' in user_input:
            parts = user_input.split('-')
            if len(parts) == 2:
                start_str = parts[0].strip()
                end_str = parts[1].strip()
                start_date = parse_date_input(start_str)
                end_date = parse_date_input(end_str)
                if start_date and end_date:
                    if start_date > end_date:
                        start_date, end_date = end_date, start_date
                    return start_date, end_date
                else:
                    print('日期格式无效，请重新输入')
                    continue
        single_date = parse_date_input(user_input)
        if single_date:
            return single_date, single_date
        else:
            print('日期格式无效，请重新输入')


def read_csv_file(csv_file_path):
    """读取CSV文件（自动识别 utf-8-sig / gbk 编码）"""
    if not os.path.exists(csv_file_path):
        print('CSV文件不存在:', csv_file_path)
        return None
    try:
        try:
            df = pd.read_csv(csv_file_path, encoding='utf-8-sig', dtype=str)
        except Exception:
            df = pd.read_csv(csv_file_path, encoding='gbk', dtype=str)
        print(' ', os.path.basename(csv_file_path), ':', len(df), '条')
        return df
    except Exception as e:
        print('读取CSV失败:', e)
        return None


# 用于逐字段比对的核心字段（与 compare_csv.py 保持一致）
FIELDS = [
    '收货单位', '收货地址', '收货人', '客户要求', '数量', '重量',
    '发运方式', '到货城市', '到货省份', '产品特性',
]


def build_dict(df, date_col='下单日期'):
    """把 DataFrame 转成 {单号: {字段: 值}} 字典"""
    result = {}
    for _, row in df.iterrows():
        order_no = str(row['单号']) if pd.notna(row.get('单号')) else ''
        if not order_no or order_no == 'nan':
            continue
        item = {'order_date': ''}
        if date_col in df.columns and pd.notna(row.get(date_col)):
            item['order_date'] = str(row[date_col])
        for f in FIELDS:
            item[f] = str(row[f]) if f in row and pd.notna(row.get(f)) else ''
        result[order_no] = item
    return result


def apply_excel_highlighting(wb):
    try:
        ws = wb['差异详情']
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        bold_font = Font(bold=True)
        diff_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and str(header).startswith('差异_'):
                diff_cols.append(col)
        for col in diff_cols:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and '\u2717' in str(cell.value):
                    cell.fill = red_fill
                    cell.font = Font(color='FFFFFF', bold=True)
                    # 高亮对应的 CSV 与 CSV2 两个原值列
                    base_col_name = str(header_replace := str(ws.cell(row=1, column=col).value)).replace('差异_', '')
                    for c in range(1, ws.max_column + 1):
                        h = ws.cell(row=1, column=c).value
                        if h and str(h) in ('CSV_' + base_col_name, 'CSV2_' + base_col_name):
                            ws.cell(row=row, column=c).fill = yellow_fill
                            ws.cell(row=row, column=c).font = bold_font
                elif cell.value and '\u2713' in str(cell.value):
                    cell.font = Font(color='00AA00', bold=True)
        for sheet_name in wb.sheetnames:
            ws_s = wb[sheet_name]
            for col in range(1, ws_s.max_column + 1):
                max_length = 0
                for row in range(1, min(ws_s.max_row + 1, 100)):
                    cell_value = ws_s.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws_s.column_dimensions[ws_s.cell(row=1, column=col).column_letter].width = min(max_length + 2, 50)
            ws_s.freeze_panes = 'A2'
    except Exception as e:
        print('高亮失败:', e)


def parse_order_date(od):
    """把下单日期字符串转成 YYYY-MM-DD，无法解析返回 None"""
    if not od:
        return None
    od = str(od).strip()
    try:
        if '/' in od:
            parts = od.split('/')
            if len(parts) == 3:
                return datetime(int(parts[0]), int(parts[1]), int(parts[2])).strftime('%Y-%m-%d')
        elif '-' in od:
            return datetime.strptime(od, '%Y-%m-%d').strftime('%Y-%m-%d')
    except Exception:
        pass
    return None


def order_diff(item1, item2):
    """判断两条订单记录是否有字段差异（用于每日统计）"""
    for f in FIELDS:
        v1 = item1.get(f, '').strip()
        v2 = item2.get(f, '').strip()
        if f in ['数量', '重量']:
            if re.sub(r'[^\d.]', '', v1) != re.sub(r'[^\d.]', '', v2):
                return True
        else:
            if clean_text(v1) != clean_text(v2):
                return True
    return False


def compare_csv_files(file1, file2, start_date=None, end_date=None):
    """两个CSV互相比对"""
    if not os.path.exists(file1):
        print('CSV文件1不存在:', file1)
        return
    if not os.path.exists(file2):
        print('CSV文件2不存在:', file2)
        return

    if start_date is None and end_date is None:
        start_date, end_date = get_date_range_from_user()

    print()
    print('文件1:', os.path.basename(file1))
    print('文件2:', os.path.basename(file2))
    if start_date:
        print('日期:', start_date, '至', end_date)
    else:
        print('日期: 全部')

    # === 读取两个CSV ===
    df1 = read_csv_file(file1)
    df2 = read_csv_file(file2)
    if df1 is None or df2 is None:
        return

    # === 检查必需列 ===
    for df, tag in [(df1, 'CSV1'), (df2, 'CSV2')]:
        for col in ['单号', '下单日期']:
            if col not in df.columns:
                print(tag, '缺少必要列:', col, '| 列名:', list(df.columns))
                return

    # === 按日期筛选 ===
    start, end = None, None
    if start_date and end_date:
        start = pd.to_datetime(start_date)
        end = pd.to_datetime(end_date)
    if start is not None:
        for df in (df1, df2):
            df['_dt'] = pd.to_datetime(df['下单日期'], errors='coerce')
            df.dropna(subset=['_dt'], inplace=True)
            df.drop([i for i in df.index if not (start <= df.loc[i, '_dt'] <= end)], inplace=True)
            del df['_dt']
            df.reset_index(drop=True, inplace=True)

    print('筛选后 CSV1:', len(df1), '条 | CSV2:', len(df2), '条')

    # === 构建字典 ===
    dict1 = build_dict(df1)
    dict2 = build_dict(df2)

    # ============ 每日统计 ============
    all_dates = set()
    for df in (df1, df2):
        df['_dt'] = pd.to_datetime(df['下单日期'], errors='coerce')
        for d in df['_dt'].dropna().dt.date.unique():
            all_dates.add(d.strftime('%Y-%m-%d'))
        del df['_dt']

    daily_stats = []
    for date_str in sorted(all_dates):
        orders1 = {oid for oid, it in dict1.items() if parse_order_date(it['order_date']) == date_str}
        orders2 = {oid for oid, it in dict2.items() if parse_order_date(it['order_date']) == date_str}

        common = orders1 & orders2
        extra1 = orders1 - orders2
        extra2 = orders2 - orders1

        diff_orders = set()
        for oid in common:
            if order_diff(dict1[oid], dict2[oid]):
                diff_orders.add(oid)

        daily_stats.append({
            '日期': date_str,
            'CSV1数量': len(orders1),
            'CSV2数量': len(orders2),
            '共同数量': len(common),
            '差异数量': len(diff_orders),
            'CSV1多出订单号': '\u3001'.join(sorted(extra1)) if extra1 else '',
            'CSV2多出订单号': '\u3001'.join(sorted(extra2)) if extra2 else '',
            '差异订单号': '\u3001'.join(sorted(diff_orders)) if diff_orders else '',
        })

    # ============ 差异详情 ============
    all_orders = set(dict1.keys()) & set(dict2.keys())
    diff_data = []
    field_diff_count = {f: 0 for f in FIELDS}

    def _order_date(oid):
        return dict1.get(oid, dict2.get(oid, {})).get('order_date', '')
    sorted_orders = sorted(all_orders, key=_order_date)

    for oid in sorted_orders:
        it1 = dict1[oid]
        it2 = dict2[oid]
        row = {'日期': it1['order_date'] or it2['order_date'], '订单号': oid}
        has_diff = False
        for f in FIELDS:
            v1 = it1.get(f, '').strip()
            v2 = it2.get(f, '').strip()
            row['CSV_' + f] = v1
            row['CSV2_' + f] = v2
            if f in ['数量', '重量']:
                n1 = re.sub(r'[^\d.]', '', v1)
                n2 = re.sub(r'[^\d.]', '', v2)
                if n1 != n2:
                    row['差异_' + f] = '\u2717'
                    field_diff_count[f] += 1
                    has_diff = True
                else:
                    row['差异_' + f] = '\u2713'
            else:
                if clean_text(v1) != clean_text(v2):
                    row['差异_' + f] = '\u2717'
                    field_diff_count[f] += 1
                    has_diff = True
                else:
                    row['差异_' + f] = '\u2713'
        if has_diff:
            diff_data.append(row)

    # ============ 输出Excel ============
    date_suffix = ''
    if start_date and end_date:
        date_suffix = ('_' + start_date) if start_date == end_date else '_' + start_date + '_至_' + end_date
    output_file = 'CSV2CSV差异报告' + date_suffix + '.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 子表1 每日统计
        df_daily = pd.DataFrame(daily_stats)
        if not df_daily.empty:
            total_row = {
                '日期': '合计',
                'CSV1数量': int(df_daily['CSV1数量'].sum()),
                'CSV2数量': int(df_daily['CSV2数量'].sum()),
                '共同数量': int(df_daily['共同数量'].sum()),
                '差异数量': int(df_daily['差异数量'].sum()),
                'CSV1多出订单号': '',
                'CSV2多出订单号': '',
                '差异订单号': '',
            }
            df_daily = pd.concat([df_daily, pd.DataFrame([total_row])], ignore_index=True)
        df_daily.to_excel(writer, sheet_name='每日统计', index=False)

        # 子表2 差异详情
        if diff_data:
            df_out = pd.DataFrame(diff_data)
            stats_row = {'日期': '', '订单号': '=== 差异统计 ==='}
            for f in FIELDS:
                stats_row['CSV_' + f] = ''
                stats_row['CSV2_' + f] = ''
                stats_row['差异_' + f] = int(field_diff_count[f]) if field_diff_count[f] > 0 else ''
            df_out = pd.concat([df_out, pd.DataFrame([stats_row])], ignore_index=True)
        else:
            df_out = pd.DataFrame([{'日期': '', '订单号': '所有订单完全一致，无差异'}])
        df_out.to_excel(writer, sheet_name='差异详情', index=False)

    try:
        wb = load_workbook(output_file)
        apply_excel_highlighting(wb)
        wb.save(output_file)
    except Exception as e:
        print('高亮失败:', e)

    # ============ 精简输出 ============
    print()
    print('已导出:', output_file)
    print('差异订单:', len(diff_data), '个 | 差异字段:', sum(field_diff_count.values()), '处')
    diff_fields = [(n, c) for n, c in field_diff_count.items() if c > 0]
    if diff_fields:
        for n, c in diff_fields:
            print(' .', n, ':', c, '处')

    print()
    print('每日统计:')
    for ds in daily_stats:
        extras = []
        if ds['CSV1多出订单号']:
            extras.append('CSV1多' + str(ds['CSV1数量'] - ds['共同数量']))
        if ds['CSV2多出订单号']:
            extras.append('CSV2多' + str(ds['CSV2数量'] - ds['共同数量']))
        extra_str = ' (' + ', '.join(extras) + ')' if extras else ''
        print('  ', ds['日期'], ': CSV1', ds['CSV1数量'], ' CSV2', ds['CSV2数量'],
              ' 共同', ds['共同数量'], ' 差异', ds['差异数量'], extra_str)


if __name__ == '__main__':
    default_dir = os.path.dirname(os.path.abspath(__file__))
    file1 = os.path.join(default_dir, 'file', '自动接单表_PPG 芜湖(工业漆)_总接单表.csv')
    file2 = os.path.join(default_dir, 'file', '自动接单表_PPG 芜湖(工业漆) 副本_总接单表.csv')

    print()
    print('=' * 60)
    print('PPG芜湖工业漆 两个CSV互相比对工具')
    print('=' * 60)
    print('说明：对比两个格式相同的CSV文件，逐订单逐字段比对差异')
    print('比对字段：' + '、'.join(FIELDS))
    print('=' * 60)

    if not os.path.exists(file1):
        print('找不到CSV文件1:', file1)
        file1 = input('请输入CSV文件1路径: ').strip()
    if not os.path.exists(file2):
        print('找不到CSV文件2:', file2)
        file2 = input('请输入CSV文件2路径: ').strip()

    compare_csv_files(file1, file2, start_date=None, end_date=None)
